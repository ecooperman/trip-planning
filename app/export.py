"""Generates a formatted .xlsx itinerary for a trip - a real spreadsheet
(colors, merged headers, sized columns), not a bare CSV, laid out as a
block per day: a shaded date header, a Time/Activity/Place/Address/Notes
table, then that day's rows. This deliberately matches the hand-maintained
format Evan already used for trip planning before this app existed (see
the reference file he shared), so an export opens exactly as readable and
familiar as the original - meant as an offline-friendly backup (the
"Export" button on trip.html/agenda.html) for when connectivity drops on
the actual trip.
"""

import re
from datetime import datetime, timedelta
from io import BytesIO
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import models

COLUMN_HEADERS = ["Time", "Activity", "Place", "Address", "Notes"]
# Matches the reference spreadsheet's own column widths - Address/Notes are
# wide since they hold full street addresses and free-text notes.
COLUMN_WIDTHS = [20.38, 18.5, 27.63, 46.88, 69.25]

HEADER_FILL = PatternFill(start_color="FFE7E6E6", end_color="FFE7E6E6", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, size=11)
HEADER_ALIGN = Alignment(horizontal="center")
BODY_FONT = Font(name="Arial", size=11)


def safe_filename(location: str) -> str:
    """A trip's location as a filesystem/header-safe .xlsx filename."""
    cleaned = re.sub(r'[\\/:*?"<>|]', "", location or "Trip").strip()
    return f"{cleaned or 'Trip'}.xlsx"


def _sheet_title(location: str) -> str:
    # Excel sheet names: 31 chars max, and the same punctuation forbidden
    # in filenames is forbidden here too.
    cleaned = re.sub(r'[\\/:*?"<>|\[\]]', "", location or "Trip").strip()
    return (cleaned or "Trip")[:31]


def _day_range(start: datetime, end: datetime) -> List:
    days = []
    cur = start.date()
    last = end.date()
    while cur <= last:
        days.append(cur)
        cur += timedelta(days=1)
    return days


def _format_time_range(start: Optional[datetime], end: Optional[datetime]) -> str:
    # %-I (no leading zero) is a glibc/macOS strftime extension, not
    # portable to Windows - fine here since this only ever runs on the
    # droplet (Linux) or a dev Mac, never deployed to Windows.
    if not start:
        return ""
    start_label = start.strftime("%-I:%M %p")
    if not end:
        return start_label
    end_label = end.strftime("%-I:%M %p")
    return f"{start_label} – {end_label}"


def _notes_for(confirmation_number: Optional[str], notes: Optional[str]) -> str:
    parts = []
    if confirmation_number:
        parts.append(f"Confirmation #: {confirmation_number}")
    if notes:
        parts.append(notes)
    return "\n".join(parts)


def _activity_row(activity: "models.Activity") -> List[str]:
    return [
        _format_time_range(activity.scheduled_start, activity.scheduled_end),
        activity.name,
        "",
        activity.address or "",
        _notes_for(activity.confirmation_number, activity.notes),
    ]


def _travel_row(segment: "models.TravelSegment") -> List[str]:
    # Mirrors an Activity row's shape (Activity = the thing's own name, not
    # a generic action word like Stay's Check-in/Check-out) - a travel
    # segment is the primary subject of its own row same as an activity is.
    # `type` is a str-backed enum (models.TravelType), so this needs no
    # separate label table - "rental_car" -> "Rental Car" the same way
    # every other type name already reads correctly this way.
    route = " → ".join(filter(None, [segment.departure_location, segment.arrival_location]))
    return [
        _format_time_range(segment.departure_time, segment.arrival_time),
        segment.name,
        segment.type.replace("_", " ").title(),
        route,
        _notes_for(segment.confirmation_number, segment.notes),
    ]


class _RowWriter:
    """Small stateful helper so the day/header/data-row writers below don't
    all have to pass the current row index back and forth by hand."""

    def __init__(self, ws):
        self.ws = ws
        self.row = 1

    def write_section_header(self, title):
        """A merged, shaded header row spanning all columns - used for a
        day's date (with real date formatting) or the trailing
        "Unscheduled" section."""
        cell = self.ws.cell(row=self.row, column=1, value=title)
        if isinstance(title, datetime):
            cell.number_format = "dddd, mmmm d"
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        self.ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=len(COLUMN_HEADERS))
        self.row += 1

    def write_column_headers(self):
        for col, value in enumerate(COLUMN_HEADERS, start=1):
            cell = self.ws.cell(row=self.row, column=col, value=value)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
        self.row += 1

    def write_data_row(self, values):
        for col, value in enumerate(values, start=1):
            cell = self.ws.cell(row=self.row, column=col, value=value)
            cell.font = BODY_FONT
        self.row += 1


def build_trip_workbook(
    trip: models.Trip,
    activities: List[models.Activity],
    stays: List[models.Stay],
    travel_segments: Optional[List[models.TravelSegment]] = None,
) -> Workbook:
    travel_segments = travel_segments or []

    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title(trip.location)
    for i, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    writer = _RowWriter(ws)

    # The outbound leg (earliest departure) and the return leg (latest
    # departure) bookend the *whole* trip, not just their own day - you
    # land before you can check in anywhere, and you check out before
    # heading to the airport, so these two get written outside the normal
    # check-in/activities/check-out order for their day rather than
    # sorted in among it (see below). Meaningless with 0-1 segments (a
    # single segment is just "the outbound," nothing to bookend the end
    # with yet), so only identified once there are at least two.
    scheduled_travel = [t for t in travel_segments if t.departure_time]
    outbound_travel = min(scheduled_travel, key=lambda t: t.departure_time) if len(scheduled_travel) >= 2 else None
    return_travel = max(scheduled_travel, key=lambda t: t.departure_time) if len(scheduled_travel) >= 2 else None

    # Activities and any other travel segment (a connecting flight, a
    # mid-trip rental car pickup, ...) both carry a real timestamp, so they
    # sort together into one chronological list per day, same as before
    # travel segments existed.
    middle_travel = [t for t in scheduled_travel if t is not outbound_travel and t is not return_travel]
    scheduled = sorted(
        [(lambda a=a: _activity_row(a), a.scheduled_start) for a in activities if a.scheduled_start]
        + [(lambda t=t: _travel_row(t), t.departure_time) for t in middle_travel],
        key=lambda entry: entry[1],
    )
    unscheduled_activities = [a for a in activities if not a.scheduled_start]
    unscheduled_travel = [t for t in travel_segments if not t.departure_time]

    if trip.start_date and trip.end_date:
        days = _day_range(trip.start_date, trip.end_date)
    else:
        # No trip dates set yet - fall back to whatever days actually have
        # something scheduled, so the export still produces something
        # useful instead of an empty sheet.
        days = sorted({start.date() for _, start in scheduled})
    # A flight can depart the day before the trip's own start_date (that's
    # often "the first day at the destination," not "the first day of
    # travel") - make sure the outbound/return day is always covered even
    # when it falls outside the declared range, so neither one is ever
    # silently dropped from the export.
    bookend_days = {t.departure_time.date() for t in (outbound_travel, return_travel) if t}
    days = sorted(set(days) | bookend_days)

    for day in days:
        writer.write_section_header(datetime(day.year, day.month, day.day))
        writer.write_column_headers()

        if outbound_travel and outbound_travel.departure_time.date() == day:
            writer.write_data_row(_travel_row(outbound_travel))

        # Stays don't carry a time-of-day (just a date), so check-in rows
        # go first and check-out rows go last for that day, with the
        # day's actual timed activities and travel segments sorted in
        # between - the same rough ordering the reference spreadsheet uses
        # by hand.
        for stay in stays:
            if stay.start_date and stay.start_date.date() == day:
                writer.write_data_row(["", "Check-in", stay.name, stay.address or "", ""])

        for row_fn, start in scheduled:
            if start.date() != day:
                continue
            writer.write_data_row(row_fn())

        for stay in stays:
            if stay.end_date and stay.end_date.date() == day:
                writer.write_data_row(["", "Check-out", stay.name, stay.address or "", ""])

        if return_travel and return_travel.departure_time.date() == day:
            writer.write_data_row(_travel_row(return_travel))

    if unscheduled_activities or unscheduled_travel:
        writer.write_section_header("Unscheduled")
        writer.write_column_headers()
        for segment in unscheduled_travel:
            writer.write_data_row(_travel_row(segment))
        for activity in unscheduled_activities:
            writer.write_data_row(_activity_row(activity))

    return wb


def workbook_bytes(wb: Workbook) -> BytesIO:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
